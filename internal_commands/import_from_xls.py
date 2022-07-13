import sys
import os

sys.path.append(os.path.abspath(os.path.dirname(os.path.dirname(__file__))))

import asyncio
import uuid

from aiofile import async_open

from openpyxl import load_workbook

from tortoise import Tortoise

from models.models import Category, Product, Manufacturer
from models.resources import filename_generator

from settings import IMG_DIR, DATABASE


async def import_products(dir_path, file_name):
    wb = load_workbook(filename=os.path.join(dir_path, file_name))
    for i, sheet in enumerate(wb.sheetnames):
        if sheet.lower == "сайт":
            continue
        img_dir = None
        for d in os.listdir(dir_path):
            if d.lower() == sheet.lower():
                img_dir = d
                break
        ws = wb.worksheets[i]
        if len(next(ws.rows)) < 7:
            continue
        category, st = await Category.get_or_create(name=sheet.split(" ")[-1])
        manufacturer, st = await Manufacturer.get_or_create(name=" ".join(sheet.split(" ")[:-1]))
        with_img = False
        for j, r in enumerate(ws.rows):
            if j == 0:
                continue
            if r[6].value and img_dir:
                try:
                    ext = ""
                    img_filename = None
                    for f in os.listdir(os.path.join(dir_path, img_dir)):
                        if r[6].value.lower() + '.jpg' == f.lower():
                            ext = "." + f.split(".")[-1].lower()
                            img_filename = f
                            break
                    if img_filename:
                        filename = uuid.uuid4().hex + ext
                        img_name = os.path.join(filename[:2], filename[2:4], filename[4:6], filename)
                        async with async_open(os.path.join(dir_path, img_dir, img_filename), 'rb') as f:
                            img = await f.read()
                        os.makedirs(os.path.join(IMG_DIR, os.path.dirname(img_name)), exist_ok=True)
                        async with async_open(os.path.join(IMG_DIR, img_name), 'wb') as f:
                            await f.write(img)
                except BaseException as e:
                    print("EXCEPTION: ", e)
                    filename = None
                product_data = {
                    "name": r[0].value,
                    "full_name": r[1].value,
                    "img": filename,
                    "description": r[4].value if r[4].value else "",
                    "sketches": r[5].value if r[5].value else "",
                    "article_number": r[2].value,
                    "volume": r[3].value,
                }
            else:
                product_data = {
                    "name": r[0].value,
                    "full_name": r[1].value,
                    "description": r[4].value if r[4].value else "",
                    "sketches": r[5].value if r[5].value else "",
                    "article_number": r[2].value,
                    "volume": r[3].value,
                }
            product = await Product.create(**product_data)
            product.category = category
            product.manufacturer = manufacturer
            await product.save()


async def main():
    dir_path = sys.argv[1]
    file_name = sys.argv[2]
    await Tortoise.init(
        db_url="postgres://{}:{}@{}:5432/{}".format(
            DATABASE["user"],
            DATABASE["password"],
            DATABASE["address"],
            DATABASE["name"],
        ),
        modules={"models": ["models.models", "aerich.models"]},
    )
    await import_products(dir_path, file_name)


if __name__ == "__main__":
    asyncio.run(main())
